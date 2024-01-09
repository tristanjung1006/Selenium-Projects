import openpyxl
import requests
from bs4 import BeautifulSoup
from selenium.common import StaleElementReferenceException
from datetime import datetime

user_agent = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 "
              "Safari/537.36")


# A열의 마지막 행을 찾는 함수
def find_last_row(sheet):
    for row in range(sheet.max_row, 1, -1):
        if sheet.cell(row=row, column=1).value is not None:
            print(f'{sheet} 링크가 {row}열까지 기입되어있습니다.')
            return row


# 기본 처리 함수
def process_sheet(sheet, custom_parsing_func):
    # A열의 마지막 행 찾기
    last_row = find_last_row(sheet)

    # URL이 저장된 열 선택 (A 열의 범위)
    url_column = sheet[f'A2:A{last_row}']

    # 상품명을 저장할 열 선택 (B 열의 범위)
    product_name_column = sheet[f'B2:B{last_row}']

    # 쿠폰적용가를 저장할 열 선택 (C 열의 범위)
    coupon_price_column = sheet[f'C2:C{last_row}']

    # 상품금액을 저장할 열 선택 (D 열의 범위)
    current_price_column = sheet[f'D2:D{last_row}']

    # 회사명을 저장할 열 선택 (E 열의 범위)
    company_column = sheet[f'E2:E{last_row}']

    # 대표 이미지 링크를 저장할 열 선택 (F 열의 범위)
    img_link_column = sheet[f'F2:F{last_row}']

    # 상세 페이지 이미지 링크를 저장할 열 선택 (G 열의 범위)
    detail_link_column = sheet[f'G2:G{last_row}']

    # URL 순회
    for index, (url_cell, product_name_cell, current_price_cell, coupon_price_cell, company_cell, img_link_cell,
                detail_link_cell) in enumerate(zip(url_column, product_name_column, current_price_column,
                                                   coupon_price_column, company_column, img_link_column,
                                                   detail_link_column), start=2):
        url = url_cell[0].value
        print(url)

        if url:
            # 현재 가격과 품절 여부 파싱 로직 함수 호출
            current_price, coupon_price, product_name, company, img_link = custom_parsing_func(
                url)  # 사용자 정의 파싱 로직 함수 호출

            # 상품명을 B 열에 저장
            product_name_cell[0].value = product_name

            # 상품금액을 C 열에 저장
            current_price_cell[0].value = current_price

            # 회사명을 D 열에 저장
            company_cell[0].value = company

            # 이미지 링크를 E 열에 저장
            current_price_cell[0].value = img_link

        else:
            print(f'URL을 찾을 수 없습니다. (행 {index})')


def try_action(action, *args, max_retries=3):
    for _ in range(max_retries):
        try:
            return action(*args)
        except StaleElementReferenceException:
            continue
    # 예외가 여전히 발생하면 처리할 로직 추가
    print("Max retries reached. Unable to perform the action.")
    return None


# 사용자 정의 파싱 로직을 포함한 함수
def gmarket_logic(url):
    # 웹페이지 내용을 가져오기
    response = requests.get(url, headers={"User-Agent": user_agent})
    html = response.text

    # BeautifulSoup을 사용하여 HTML 파싱
    soup = BeautifulSoup(html, 'html.parser')

    # price_real 클래스를 가진 strong 태그 추출
    price_real_strong = soup.find('strong', class_='price_real')
    if price_real_strong:
        price_real_value = price_real_strong.text
        print(f"price_real 클래스를 가진 strong 태그의 텍스트 값: {price_real_value}")
    else:
        print("price_real 클래스를 가진 strong 태그를 찾을 수 없습니다.")

    # price_innerwrap, price_innerwrap-coupon 클래스를 가진 span 태그의 하위에 있는 price_real 클래스를 가진 strong 태그 추출
    price_innerwrap_span = soup.find('span', class_='price_innerwrap price_innerwrap-coupon')
    if price_innerwrap_span:
        nested_price_real_strong = price_innerwrap_span.find('strong', class_='price_real')
        if nested_price_real_strong:
            nested_price_real_value = nested_price_real_strong.text
            print(f"span 태그의 하위에 있는 price_real 클래스를 가진 strong 태그의 텍스트 값: {nested_price_real_value}")
        else:
            print("span 태그의 하위에 있는 price_real 클래스를 가진 strong 태그를 찾을 수 없습니다.")
    else:
        print("price_innerwrap, price_innerwrap-coupon 클래스를 가진 span 태그를 찾을 수 없습니다.")

    # img_link = soup.find('div', class_='thumb-gallery uxecarousel')

    img_element = soup.find('img')

    # img 태그의 src 속성 가져오기
    img_src = img_element.get('src')

    img_link = 'https:' + img_src
    print(img_link)

    # 이미지 다운로드
    if img_link:
        # 현재 시간을 기반으로 파일 이름 생성
        current_time = datetime.now().strftime("%Y%m%d%H%M%S")
        file_name = f'downloaded_image_{current_time}.jpg'

        response = requests.get(img_link, headers={"User-Agent": user_agent})
        if response.status_code == 200:
            # 이미지 다운로드
            with open(file_name, 'wb') as f:
                f.write(response.content)
            print(f'이미지 다운로드 완료: {file_name}')
        else:
            print('이미지 다운로드 실패. 상태 코드:', response.status_code)

    # ee-image 클래스를 가진 div 태그를 찾음
    div_tag = soup.find('div', class_='ee-image')

    # div 태그가 존재하면서 img 태그를 찾음
    if div_tag:
        img_tag = div_tag.find('img')

        # img 태그가 존재하면서 src 속성을 추출
        if img_tag:
            src_value = img_tag.get('src')
            print(f"추출된 사이트 주소: {src_value}")
        else:
            print("img 태그를 찾을 수 없습니다.")
    else:
        print("ee-image 클래스를 가진 div 태그를 찾을 수 없습니다.")

    return img_link


# 시트 이름과 사용자 정의 파싱 함수 매핑
sheet_custom_parsing_mapping = {
    'gmarket': gmarket_logic
}

for iteration in range(1):
    print(f"{iteration + 1}번째 반복 시작")

    # 엑셀 파일 열기
    workbook = openpyxl.load_workbook('product_link.xlsx')

    # 모든 시트를 순회하면서 함수 호출
    for sheet_name, custom_parsing_func in sheet_custom_parsing_mapping.items():
        current_sheet = workbook[sheet_name]
        process_sheet(current_sheet, custom_parsing_func)

    # 엑셀 파일 저장
    workbook.save('product_link.xlsx')

    print(f"{iteration + 1}번째 반복 완료 및 엑셀 파일 저장")
