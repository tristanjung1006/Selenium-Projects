import time

import openpyxl
import requests
from bs4 import BeautifulSoup
import json
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl.styles import PatternFill
from selenium.webdriver.chrome.options import Options
from urllib.parse import urlparse, parse_qs
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException, TimeoutException, ElementClickInterceptedException, UnexpectedAlertPresentException

user_agent = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 "
              "Safari/537.36")

# 웹브라우저 창 크기 및 웹페이지 요소 로딩 대기시간 설정
service = Service(ChromeDriverManager().install())
chrome_options = Options()
chrome_options.add_argument("--headless")
chrome_options.add_argument('--window-size=1920,1080')
chrome_options.add_argument(f"user-agent={user_agent}")
chrome_options.add_argument("--allow-running-insecure-content")
chrome_options.add_argument("--log-level=3")
chrome_options.add_argument('--disable-logging')
driver = webdriver.Chrome(service=service, options=chrome_options)
driver.maximize_window()
wait = WebDriverWait(driver, 15)

repeat = int(input("반복횟수: "))


# A열의 마지막 행을 찾는 함수
def find_last_row(sheet):
    for row in range(sheet.max_row, 1, -1):
        if sheet.cell(row=row, column=1).value is not None:
            print(f'{sheet} 링크가 {row}열까지 기입되어있습니다.')
            return row


# 기본 처리 함수
def process_sheet(sheet, custom_parsing_func):
    # A열의 마지막 행 찾기
    last_row = find_last_row(sheet)

    # URL이 저장된 열 선택 (A 열의 범위)
    url_column = sheet[f'A2:A{last_row}']

    # 현재 가격을 저장할 열 선택 (B 열의 범위)
    current_price_column = sheet[f'B2:B{last_row}']

    # 변동 가격을 저장할 열 선택 (C 열의 범위)
    price_change_column = sheet[f'C2:C{last_row}']

    # 사이즈 옵션을 저장할 열 선택 (D 열의 범위)
    size_column = sheet[f'D2:D{last_row}']

    # 변동 가격을 저장할 열 선택 (E 열의 범위)
    soldout_column = sheet[f'E2:E{last_row}']

    # URL 순회
    for index, (url_cell, current_price_cell, price_change_cell, size_cell, soldout_cell) in enumerate(
            zip(url_column, current_price_column, price_change_column, size_column, soldout_column), start=2):
        url = url_cell[0].value
        print(url)

        if url:
            # 기존 변동 가격을 현재 가격으로 설정하기
            price_change_cell[0].value = current_price_cell[0].value

            # 현재 가격과 품절 여부 파싱 로직 함수 호출
            current_price, soldout = custom_parsing_func(url, size_cell[0].value)  # 사용자 정의 파싱 로직 함수 호출

            # 현재 가격을 B 열에 저장
            current_price_cell[0].value = current_price

            # 품절 여부를 E 열에 저장
            soldout_cell[0].value = soldout

            # 기존 가격과 현재 가격 비교
            if price_change_cell[0].value != current_price:
                # 노란색으로 색칠
                current_price_cell[0].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            elif current_price == "":
                # 가격값이 없을 경우 색상 없애기
                current_price_cell[0].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            else:
                # 값이 같을 경우 색상 없애기
                current_price_cell[0].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        else:
            print(f'URL을 찾을 수 없습니다. (행 {index})')


def try_action(action, *args, max_retries=3):
    for _ in range(max_retries):
        try:
            return action(*args)
        except StaleElementReferenceException:
            continue
    # 예외가 여전히 발생하면 처리할 로직 추가
    print("Max retries reached. Unable to perform the action.")
    return None


# 사용자 정의 파싱 로직을 포함한 함수
def okmall_logic(url, size_cell):
    # 웹페이지 내용을 가져오기
    response = requests.get(url, headers={"User-Agent": user_agent})
    html = response.text

    # BeautifulSoup을 사용하여 HTML 파싱
    soup = BeautifulSoup(html, 'html.parser')
    # 'last_price' 클래스를 가진 div 태그 찾기
    last_price_element = soup.find('div', class_='last_price')

    if last_price_element:
        # 'last_price' 클래스를 가진 div 태그 내의 모든 span 태그를 찾고, 두 번째 span 태그의 텍스트 추출
        span_elements = last_price_element.find_all('span')
        if len(span_elements) > 1:
            current_price = span_elements[2].get_text(strip=True)
        else:
            print('두 번째 span 태그를 찾을 수 없습니다.')
            current_price = ""
    else:
        print('가격 정보를 찾을 수 없습니다.')
        current_price = ""

    # 'btn_alarm' 클래스를 가진 span 태그 찾기
    btn_alarm_element = soup.find('span', class_='btn_alarm')

    # 'btn_alarm' 클래스를 가진 span 태그가 존재하는지 확인
    if btn_alarm_element:
        soldout = "품절"
    else:
        soldout = ""

    # 't_size' 클래스를 가진 span 태그 찾기
    size_element = soup.find('span', class_='t_size')

    if size_element:
        # 't_size' 클래스를 가진 span태그에서 텍스트를 추출
        size = size_element.get_text(strip=True)
        try:
            if str(size_cell) in size:
                print("해당하는 사이즈의 제품은 재고가 있습니다.")
                soldout = ""
            else:
                print("해당하는 사이즈의 제품은 재고가 없습니다.")
                soldout = "품절"
        except AttributeError:
            print("엑셀파일에 사이즈 옵션을 설정하지 않은 제품입니다.")
            pass

    # 다른 파싱 작업을 수행하고 필요한 현재 가격과 변동 가격을 반환
    # 현재 가격 및 변동 가격을 예시로 추출한 것입니다.
    print("현재 가격: " + current_price + " " + soldout)
    return current_price, soldout


# 옵션 품절 여부(셀레니움)
def trendmecca_logic(url, size_cell):
    # 웹페이지 내용을 가져오기
    response = requests.get(url, headers={"User-Agent": user_agent})
    html = response.text

    # BeautifulSoup을 사용하여 HTML 파싱
    soup = BeautifulSoup(html, 'html.parser')

    try:
        optimum_discount_price = soup.find('span', id='span_optimum_discount_price')
        price_text = ''

        for element in optimum_discount_price.contents:
            if element.name == 'span' and 'style' in element.attrs:
                # 스타일 속성을 가진 span 태그 (무시)
                continue
            # 스타일이 없는 태그의 텍스트 추출
            price_text += str(element)

        # 'oban_list' 클래스를 가진 ol 태그 찾기
        ol_element = soup.find('ol', class_='oban_list')

        # 결과를 저장할 리스트 초기화
        strong_texts = []

        if ol_element:
            # ol 태그의 하위 li 태그들 중에서 soldout 클래스를 가진 것들 찾기
            soldout_li_elements = ol_element.find_all('li', class_='soldout')

            for soldout_li in soldout_li_elements:
                # li 태그의 하위 a 태그 찾기
                a_element = soldout_li.find('a')
                if a_element:
                    # a 태그의 하위 div 태그 중 oban_model 클래스를 가진 것 찾기
                    oban_model_div = a_element.find('div', class_='oban_model')
                    if oban_model_div:
                        # div 태그의 하위 strong 태그 텍스트 추출하여 리스트에 저장
                        strong_element = oban_model_div.find('strong')
                        if strong_element:
                            strong_texts.append(strong_element.get_text(strip=True))
                            if str(size_cell) in strong_texts:
                                soldout = "품절"
                            else:
                                soldout = ""
                        else:
                            print("Strong 태그를 찾을 수 없습니다.")
                            soldout = ""
                    else:
                        print("oban_model 클래스를 가진 div 태그를 찾을 수 없습니다.")
                        soldout = ""
                else:
                    print("a 태그를 찾을 수 없습니다.")
                    soldout = ""
        else:
            print("본 제품은 단품 상품입니다.")
            soldout = ""

    except AttributeError:
        price_text = ""
        soldout = "품절"

    # 결과 출력
    current_price = price_text.strip()
    print(f'현재 가격: {current_price} {soldout}')
    return current_price, soldout


def folder_logic(url, size_cell):
    # 웹페이지 내용을 가져오기
    response = requests.get(url, headers={"User-Agent": user_agent})
    html = response.text

    # BeautifulSoup을 사용하여 HTML 파싱
    soup = BeautifulSoup(html, 'html.parser')

    # 'headquarters--discount' 클래스를 가진 span 태그 찾기
    span_element = soup.find('span', class_='headquarters--discount')

    if span_element:
        # span 태그의 하위 em 태그 찾기
        em_element = span_element.find('em')

        if em_element:
            # em 태그 안의 텍스트 추출
            text_inside_em = em_element.get_text(strip=True)
            current_price = text_inside_em
        else:
            print("em 태그를 찾을 수 없습니다.")
            current_price = ""
            soldout = ""
    else:
        print("headquarters--discount 클래스를 가진 span 태그를 찾을 수 없습니다.")
        current_price = ""
        soldout = ""

    # JSON 파싱을 위한 url 변경
    url = url.replace("shop/goodsView", "controller/product/loadOptionDatas")
    # 웹페이지 내용 가져오기
    response = requests.get(url, headers={"User-Agent": user_agent})

    if response.status_code == 200:
        data = response.text
        start_index = data.find("var devOptionData = {")  # devOptionData 객체 시작 위치 찾기
        if start_index != -1:
            data = data[start_index:]
            end_index = data.find("};")  # devOptionData 객체 끝 위치 찾기
            if end_index != -1:
                data = data[:end_index + 1]

                # JSON 형식으로 파싱
                devOptionData = json.loads(data[len("var devOptionData = "):])
                viewOptions = devOptionData.get("viewOptions", [])

                for option in viewOptions:
                    if option.get("option_name") == "사이즈":
                        optionDetailList = option.get("optionDetailList", [])
                        for detail in optionDetailList:
                            if detail.get("option_div") == str(size_cell):
                                option_stock = detail.get("option_stock")
                                if option_stock == "0":
                                    soldout = "품절"
                                    print(f"{size_cell} 사이즈 재고량: {option_stock}")
                                else:
                                    soldout = ""
                                    print(f"{size_cell} 사이즈 재고량: {option_stock}")
                                break
                            else:
                                soldout = ""
                    else:
                        soldout = ""
            else:
                print("devOptionData 객체의 끝을 찾을 수 없습니다.")
                soldout = ""
        else:
            print("devOptionData 객체의 시작을 찾을 수 없습니다.")
            soldout = ""
    else:
        print(f"HTTP 요청 오류: {response.status_code}")
        soldout = ""

    print("현재 가격: " + current_price + " " + soldout)
    return current_price, soldout


def musinsa_logic(url, size_cell):
    # 웹페이지 내용을 가져오기
    response = requests.get(url, headers={"User-Agent": user_agent})
    html = response.text

    # BeautifulSoup을 사용하여 HTML 파싱
    soup = BeautifulSoup(html, 'html.parser')

    # product_article_price라는 class를 가진 span 태그 찾기
    price_element = soup.find('span', class_='product_article_price')
    if price_element:
        current_price = price_element.get_text(strip=True)
    else:
        current_price = ""

    # option1이라는 id를 가진 select 태그 찾기
    option_element = soup.find('select', id='option1')

    # option_element가 존재하는 경우에만 계속 진행
    if option_element:
        # option_element 하위에 있는 모든 option 태그 찾기
        option_tags = option_element.find_all('option')

        # 각 option 태그의 텍스트값을 리스트에 저장
        option_texts = [option.text.strip() for option in option_tags]
        print(option_texts)

        for option in option_texts:
            if str(size_cell) in option:
                if "품절" in option:
                    soldout = "품절"
                    break
                else:
                    soldout = ""
            else:
                soldout = ""
    else:
        print("Option1 select 태그를 찾을 수 없습니다.")
        soldout = ""

    print("현재 가격: " + current_price + " " + soldout)
    return current_price, soldout


def shoemarker_logic(url, size_cell):
    try:
        response = requests.get(url, headers={"User-Agent": user_agent})
        html = response.text

        # BeautifulSoup을 사용하여 HTML 파싱
        soup = BeautifulSoup(html, 'html.parser')

        # price_red라는 class를 가진 span 태그 선택
        price_red_element = soup.find('span', class_='price')

        if price_red_element:
            strong_tag = price_red_element.find('strong')
            price_present = strong_tag.get_text(strip=True)
            current_price = price_present
        else:
            current_price = ""

        driver.get(url)
        # size-list 클래스를 가진 ul 태그 찾기
        size_list_element = driver.find_element(By.CLASS_NAME, "size-list")

        # chk-size 클래스를 가진 label 태그들 찾기
        chk_size_labels = size_list_element.find_elements(By.CLASS_NAME, "chk-size")

        # 텍스트 값을 추출하여 리스트에 저장
        label_texts = [label.text for label in chk_size_labels]

        # 리스트 출력
        print(label_texts)

        if str(size_cell) in label_texts:
            soldout = ""
        else:
            soldout = "품절"

        print("현재 가격: " + current_price + soldout)
        return current_price, soldout
    except UnexpectedAlertPresentException:
        print("판매중인 상품이 아닙니다.")
        current_price = ""
        soldout = ""
        return current_price, soldout


# 전부 다(셀레니움)
def lotteon_logic(url, size_cell):
    # ,을 기준으로 문자열 분리
    split_values = str(size_cell).split(',')

    # 앞 뒤에 있는 단어를 각각 다른 변수에 저장
    first_word = split_values[0].strip() if len(split_values) > 0 else ""
    second_word = split_values[1].strip() if len(split_values) > 1 else ""

    # URL 파싱
    parsed_url = urlparse(url)
    # 원하는 부분 추출
    product_id = parsed_url.path.split('/')[-1]

    if product_id[0:2] == "LE":
        # 쿼리 스트링 파싱
        query_params = parse_qs(parsed_url.query)
        # sitmNo 값 추출
        sitmNo_value = query_params.get('sitmNo', [None])[0]
        new_url = f"https://pbf.lotteon.com/product/v2/detail/search/base/sitm/{sitmNo_value}"
    else:
        new_url = f"https://pbf.lotteon.com/product/v2/detail/search/base/pd/{product_id}"

    # 새로운 URL로 요청을 보내어 JSON 데이터 가져오기
    response = requests.get(new_url, headers={"User-Agent": user_agent})
    data = response.text

    # options에서 value값이 240인 부분에서 quantity값 추출
    datas = json.loads(data)
    current_price = datas["data"]["priceInfo"]["slPrc"]
    viewOptions = datas["data"]["optionInfo"]["optionList"]
    if len(viewOptions) == 1:
        for option in viewOptions:
            for opt in option["options"]:
                if str(size_cell) in opt["label"]:
                    if str(opt["disabled"]) == "True":
                        soldout = "품절"
                    else:
                        soldout = ""
                    break
                else:
                    if str(opt["disabled"]) == "True":
                        soldout = "품절"
                    else:
                        soldout = ""
    else:
        driver.get(url)

        try:
            script = "arguments[0].scrollIntoView();"
            # 첫 번째 div 클릭
            first_option_wrap = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//div[@class='optionWrap block withLabel']"))
            )
            driver.execute_script(script, first_option_wrap)
            first_option_wrap.click()

            # first_sel의 하위 selectLists 클래스를 가진 ul 태그 선택
            ul_element = first_option_wrap.find_element(By.CLASS_NAME, "selectLists")

            # ul 태그의 하위 li 태그들 중에서 disabled 클래스를 갖고 있지 않은 것 선택
            li_elements = ul_element.find_elements(By.CSS_SELECTOR, "li:not(.disabled)")

            # 각 li 태그에서 caption 클래스를 가진 span 태그의 텍스트를 추출하여 리스트에 저장
            captions_0 = [li.text for li in li_elements]

            # 저장된 리스트 출력
            print(captions_0)
            if first_word in captions_0:
                # 첫 번째 div 하위 태그 중 selectLists 클래스를 가진 ul 태그의 하위 태그들 중 그레이(194) 텍스트를 가진 span 클릭
                gray_span_xpath = f"//ul[@class='selectLists']//span[text()='{first_word}']"
                gray_span = wait.until(EC.element_to_be_clickable((By.XPATH, gray_span_xpath)))
                gray_span.click()

                sel = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "topOptionContent")))
                first_sel = sel.find_elements(By.XPATH,
                                              ".//*[contains(@class, 'optionWrap') and contains(@class, 'block') and contains("
                                              "@class, 'withLabel')]")[1]
                first_sel.click()

                # first_sel의 하위 selectLists 클래스를 가진 ul 태그 선택
                ul_element = first_sel.find_element(By.CLASS_NAME, "selectLists")

                # ul 태그의 하위 li 태그들 중에서 disabled 클래스를 갖고 있지 않은 것 선택
                li_elements = ul_element.find_elements(By.CSS_SELECTOR, "li:not(.disabled)")

                # 각 li 태그에서 caption 클래스를 가진 span 태그의 텍스트를 추출하여 리스트에 저장
                captions = [li.find_element(By.CLASS_NAME, "caption").text for li in li_elements]

                # 저장된 리스트 출력
                print(captions)
                if second_word in captions:
                    soldout = ""
                else:
                    soldout = "품절"
            else:
                soldout = "품절"
        except TimeoutException:
            # 첫 번째 buttonWrap 클래스를 가진 div 태그 찾기
            first_button_wrap_div = driver.find_element(By.CLASS_NAME, "buttonWrap")

            # 두 번째 buttonWrap 클래스를 가진 div 태그 찾기
            second_button_wrap_div = driver.find_element(By.XPATH, "(//div[@class='buttonWrap'])[2]")

            # 첫 번째 div 태그의 하위에 있는 labelText 클래스를 가진 span 태그들 찾기
            first_label_text_spans = first_button_wrap_div.find_elements(By.CLASS_NAME, "labelText")

            # 두 번째 div 태그의 하위에 있는 labelText 클래스를 가진 span 태그들 찾기
            second_label_text_spans = second_button_wrap_div.find_elements(By.CLASS_NAME, "labelText")

            # 텍스트 추출 및 리스트에 저장
            first_label_text_values = [span.text for span in first_label_text_spans]
            second_label_text_values = [span.text for span in second_label_text_spans]

            # 결과 출력
            print(first_label_text_values)
            print(second_label_text_values)

            if first_word in first_label_text_values or first_word in second_label_text_values:
                soldout = ""
            else:
                soldout = "품절"
        except ElementClickInterceptedException:
            soldout = ""

    print(f'현재 가격: {current_price} {soldout}')
    return current_price, soldout


def hmall_logic(url, size_cell):
    # "slitmCd=" 다음의 인덱스를 찾기
    start_index = url.find("slitmCd=")

    # 숫자 시작 인덱스 계산하여 10자리 숫자 추출
    number_start_index = start_index + len("slitmCd=")
    slitmCd_number = url[number_start_index:number_start_index + 10]

    price_url = f"https://www.hmall.com/api/hf/dp/v1/item-ptc/item-basic?slitmCd={slitmCd_number}"
    response = requests.get(price_url, headers={"User-Agent": user_agent})
    data = response.text
    datas = json.loads(data)
    current_price = datas["respData"]["itemPtc"]["bbprc"]

    option_url = f"https://www.hmall.com/api/hf/dp/v1/item-ptc/item-stockcount?slitmCd={slitmCd_number}"
    response = requests.get(option_url, headers={"User-Agent": user_agent})
    data_opt = response.text
    datas_opt = json.loads(data_opt)
    options = datas_opt["respData"]["stockList"]
    quantity = ""
    for option in options:
        if str(size_cell) in option.get("uitmTotNm"):
            quantity = option["stockCount"]
            soldout_opt = option["sellGbcd"]
            if soldout_opt == "00":
                soldout = ""
                print(f"{str(size_cell)}: 해당 옵션의 재고가 있습니다.")
            else:
                soldout = "품절"
                print(f"{str(size_cell)}: 해당 옵션의 재고가 없습니다.")
            break
        else:
            soldout = "품절"

    # 결과 출력
    print(f'현재 가격: {current_price} 재고량: {quantity} {soldout}')
    return current_price, soldout


# 옵션품절 여부(셀레니움)
def ssg_logic(url, size_cell):
    # ,을 기준으로 문자열 분리
    split_values = str(size_cell).split(',')

    # 앞 뒤에 있는 단어를 각각 다른 변수에 저장
    first_word = split_values[0].strip() if len(split_values) > 0 else ""
    second_word = split_values[1].strip() if len(split_values) > 1 else ""
    print(first_word)
    print(second_word)
    soldout = ""

    driver.get(url)
    # ssg_price 클래스를 가진 em 태그의 텍스트 추출
    ssg_price_element = wait.until(EC.presence_of_element_located((By.CLASS_NAME, "ssg_price")))
    current_price = ssg_price_element.text

    # cdtl_opt clickable 클래스를 가진 div 태그 클릭
    cdtl_opt_div = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "cdtl_opt.clickable")))
    cdtl_opt_div.click()

    # 클릭한 div 태그의 하위에 있는 txt 클래스를 가진 span 태그의 텍스트 추출
    txt_span_elements = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".cdtl_opt.clickable .txt")))
    txt_values = [element.text for element in txt_span_elements]
    print(f"옵션: {txt_values}")
    for info in txt_values:
        if first_word in info:
            if "품절" in info or "매진" in info:
                print(f"{first_word}: 해당옵션은 품절입니다.")
                soldout = "품절"
            else:
                print(f"{first_word}은(는) 품절이 아닙니다.")
                soldout = ""
            break
        else:
            soldout = ""

    # "DGY(다크그레이)"가 포함된 텍스트가 있는지 확인하고 해당하는 경우에만 처리
    if second_word != "":
        dark_gray_li = wait.until(EC.element_to_be_clickable((By.XPATH, f"//ul[@class='cdtl_select_lst _drop_list']//li[.//span[text()='{first_word}']]")))
        dark_gray_li.click()
        # 두 번째 cdtl_opt clickable 클래스를 가진 div 태그 클릭
        second_cdtl_opt_div = wait.until(
            EC.element_to_be_clickable((By.XPATH, "(//div[@class='cdtl_opt clickable'])[2]")))
        second_cdtl_opt_div.click()

        # 클릭한 div 태그의 하위에 있는 txt 클래스를 가진 span 태그의 텍스트 추출
        second_cdtl_opt_txt_spans = second_cdtl_opt_div.find_elements(By.CLASS_NAME, "txt")

        second_txt_values = [element.text for element in second_cdtl_opt_txt_spans]
        print(f"두 번째 옵션: {second_txt_values}")

        for info in second_txt_values:
            if second_word in info:
                if "품절" in info or "매진" in info:
                    print(f"{second_word}: 해당옵션은 품절입니다.")
                    soldout = "품절"
                else:
                    print(f"{second_word}은(는) 품절이 아닙니다.")
                    soldout = ""
                break
            else:
                soldout = ""

    print(f"현재 가격: {current_price} {soldout}")
    return current_price, soldout


# 현재 가격(셀레니움)
def abcmart_logic(url, size_cell):
    # 페이지 열기
    driver.get(url)

    # strong_tag = wait.until(EC.presence_of_element_located((By.XPATH, './/span[contains(@class, "price-cost")]')))
    strong_tag = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '.price-cost[data-product="sell-price'
                                                                             '-amount"]')))
    current_price = strong_tag.text
    print(current_price)

    # div 태그의 하위에 있는 모든 li 태그를 선택
    button_tags = driver.find_elements(By.XPATH, './/button[contains(@class, "btn-prod-size")]')
    button_texts = []

    if button_tags:
        for button in button_tags:
            button_texts.append(button.text)

        print(button_texts)

        # 특정 사이즈가 품절인지 확인
        if str(size_cell) in button_texts:
            soldout = ""
        else:
            soldout = "품절"
    else:
        # option = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[@id='selectboxoptionlist-button']")))
        # print(option)
        # option.click()

        option_tags = driver.find_elements(By.XPATH, './/option[contains(@data-product-type, "option")]')

        for button in option_tags:
            button_texts.append(button.text)

        # 특정 사이즈가 품절인지 확인
        if str(size_cell) in button_texts:
            soldout = ""
        else:
            soldout = ""

    print(f'현재 가격: {current_price} 재고량: {soldout}')
    return current_price, soldout


def balaan_logic(url, size_cell):
    response = requests.get(url, headers={"User-Agent": user_agent})
    html = response.text

    # BeautifulSoup을 사용하여 HTML 파싱
    soup = BeautifulSoup(html, 'html.parser')

    # total-price-view view_price member-price-2 클래스를 가진 span 태그의 텍스트 추출
    current_price = soup.find('span', class_='total-price-view view_price member-price-2').get_text(strip=True)

    # '그레이 / L' 값을 가진 option 태그의 data-stock 속성 값을 출력
    gray_l_option = soup.find('select', id='opt').find('option', {'value': {size_cell}})
    if gray_l_option:
        data_stock_value = gray_l_option.get('data-stock')
        if data_stock_value == 0:
            soldout = "품절"
        else:
            soldout = data_stock_value
    else:
        print("Option을 찾을 수 없습니다.")
        soldout = ""

    print(f'현재 가격: {current_price} {size_cell} 재고량: {soldout}')
    return current_price, soldout


def mustit_logic(url, size_cell):
    response = requests.get(url, headers={"User-Agent": user_agent})
    html = response.text

    # BeautifulSoup을 사용하여 HTML을 파싱합니다.
    soup = BeautifulSoup(html, 'html.parser')

    # 'product_price_area__bs_price' 클래스를 가진 span 태그의 하위에 있는 'price' 클래스를 가진 span 태그의 텍스트를 추출
    price_container = soup.find('span', class_='product_price_area__bs_price')
    if price_container:
        price_span = price_container.find('span', class_='price')
        if price_span:
            current_price = price_span.get_text(strip=True)
        else:
            print("'price' 클래스를 가진 span 태그를 찾을 수 없습니다.")
            current_price = ""
    else:
        print("'product_price_area__bs_price' 클래스를 가진 span 태그를 찾을 수 없습니다.")
        current_price = ""

    # 'productOption' id를 가진 div 태그의 하위에 있는 'mi-pointer' class를 가진 li 태그들 중
    # 두 번째 span 태그의 텍스트를 추출하여 리스트에 저장
    # 두 번째 span 태그의 텍스트를 추출하여 리스트에 저장
    product_option_div = soup.find('div', id='productOption')
    if product_option_div:
        second_span_texts = [li.find_all('span')[1].get_text(strip=True) for li in
                             product_option_div.find_all('li', class_='mi-pointer')]
        # size_cell의 값이 리스트에 포함되는지 확인하고 soldout 값을 설정
        if str(size_cell) in second_span_texts:
            soldout = ""
        else:
            soldout = "품절"
    else:
        print("ID가 'productOption'인 div 태그를 찾을 수 없습니다.")

    print(f'현재 가격: {current_price} {soldout}')
    return current_price, soldout


def trenbe_logic(url, size_cell):
    # URL을 '/'로 분할하고, 그 중에서 마지막 부분을 선택한 후 '+'를 기준으로 다시 분할
    url_parts = url.split('/')
    last_part = url_parts[-1].split('?')[0]

    # 추출된 부분에서 처음 8자리를 선택
    extracted_number = last_part[-8:]

    # 새로운 URL 생성
    new_url = f"https://displaygateway.trenbe.com/v1/sdp?goodsno={extracted_number}"

    # 새로운 URL로 요청을 보내어 JSON 데이터 가져오기
    response = requests.get(new_url, headers={"User-Agent": user_agent})
    data = response.text

    # options에서 value값이 240인 부분에서 quantity값 추출
    options = json.loads(data)
    current_price = options["data"]["product"]["finalPrice"]
    viewOptions = options["data"]["product"]["options"]
    quantity = ""
    for option in viewOptions:
        if str(size_cell) in option.get("value"):
            quantity = option["quantity"]
            if quantity == 0:
                soldout = "품절"
            else:
                soldout = ""
            break
        else:
            soldout = ""

    # 결과 출력
    print(f'현재 가격: {current_price} 재고량: {quantity} {soldout}')
    return current_price, soldout


def smarket_logic(url, size_cell):
    # 페이지 열기
    driver.get(url)

    strong_tag = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '.item_price_wrap strong')))
    current_price = strong_tag.find_element(By.TAG_NAME, 'em').text

    # class가 'item_size_more'인 div 태그를 선택
    item_size_more_div = wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'item_size_more')))

    # div 태그의 하위에 있는 모든 li 태그를 선택
    take_li_tags = item_size_more_div.find_elements(By.XPATH, './/li[contains(@class, "take")]')

    # li 태그의 텍스트를 추출하여 리스트에 저장
    text_list = [li.text for li in take_li_tags]
    print(text_list)

    # 특정 사이즈가 품절인지 확인
    if str(size_cell) in text_list:
        soldout = ""
    else:
        soldout = "품절"

    # 결과 출력
    print(f'현재 가격: {current_price} {soldout}')
    return current_price, soldout


def akmall_logic(url, size_cell):
    response = requests.get(url, headers={"User-Agent": user_agent})
    html = response.text

    # BeautifulSoup을 사용하여 HTML을 파싱
    soup = BeautifulSoup(html, 'html.parser')

    # class가 'item_price_wrap'인 div 태그의 하위 strong 태그를 선택
    i_tag = soup.select_one('.c_price')

    # strong 태그의 하위 em 태그의 텍스트를 추출
    if i_tag:
        current_price = i_tag.find('i').get_text()
        print(current_price)
    else:
        print("해당 요소를 찾을 수 없습니다.")
        current_price = ""

    # id가 'selectMulti'인 select 태그를 선택
    select_tag = soup.find('select', id='selectMulti')
    if not select_tag:
        select_tag = soup.find('select', id='select1')

    if select_tag:
        # select 태그의 하위에 있는 모든 옵션 태그를 선택
        option_tags = select_tag.find_all('option')

        # '품절'이라는 텍스트가 없는 옵션의 텍스트를 추출하여 리스트에 저장
        text_list = [option.get_text() for option in option_tags if '품절' not in option.get_text()]

        print(text_list)

        if str(size_cell) in text_list:
            soldout = ""
        else:
            soldout = "품절"
    else:
        print("본 제품은 단품 상품입니다.")
        soldout = ""

    print(f'현재 가격: {current_price} {soldout}')
    return current_price, soldout


def onthespot_logic(url, size_cell):
    # /product와 ? 사이에 /info 문자열을 추가하여 원하는 주소로 변환
    desired_url = url.replace("/product?", "/product/info?")

    # 변환된 주소로 HTTP GET 요청
    response = requests.get(desired_url, headers={"User-Agent": user_agent})

    # HTTP 요청이 성공했는지 확인
    if response.status_code == 200:
        # JSON 데이터 파싱
        product_data = response.json()

        try:
            # productPrice의 sellAmt 값 출력
            current_price = product_data.get("productPrice", {}).get("sellAmt")

            # productOption에서 optnName이 220인 부분의 totalStockQty 값 출력
            product_options = product_data.get("productOption", [])
            for option in product_options:
                if option.get("optnName") == str(size_cell):
                    orderPsbltQty = option.get("orderPsbltQty")
                    if orderPsbltQty == 0:
                        soldout = "품절"
                    else:
                        soldout = ""
                    break
                else:
                    orderPsbltQty = option.get("orderPsbltQty")
                    if orderPsbltQty == 0:
                        soldout = "품절"
                    else:
                        soldout = ""
        except AttributeError:
            current_price = ""
            soldout = ""
            orderPsbltQty = ""
    else:
        print(f"HTTP 요청 오류: {response.status_code}")

    print(f'현재 가격: {current_price} 재고량: {orderPsbltQty} {soldout}')
    return current_price, soldout


# 사이즈 셀레니움
def wconcept_logic(url, size_cell):
    # ,을 기준으로 문자열 분리
    split_values = str(size_cell).split(',')

    # 앞 뒤에 있는 단어를 각각 다른 변수에 저장
    first_word = split_values[0].strip() if len(split_values) > 0 else ""
    second_word = split_values[1].strip() if len(split_values) > 1 else ""

    # HTTP GET 요청
    response = requests.get(url, headers={"User-Agent": user_agent})

    # BeautifulSoup을 사용하여 HTML 파싱
    soup = BeautifulSoup(response.text, "html.parser")

    # sale 클래스를 가진 dd 태그의 하위 em 태그의 텍스트 출력
    sale_element = soup.find("dd", class_="sale")
    em_element = sale_element.find("em")
    current_price = em_element.get_text(strip=True)

    if first_word != "None":
        driver.get(url)
        first_select_list_box = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='select-list-box ']")))
        first_select_list_box.click()
        ul_element = driver.find_elements(By.CLASS_NAME, "select-list-box ")[0]
        non_soldout_texts = [a.text for a in ul_element.find_elements(By.TAG_NAME, "a") if "품절" not in a.text]
        if first_word in non_soldout_texts:
            soldout = ""
            if second_word != "":
                black_span = wait.until(
                    EC.element_to_be_clickable((By.XPATH, f"//ul[@class='select-list']//a[text()='{first_word}']")))
                black_span.click()
                # select-list-box 클래스를 가진 모든 div 태그 가져오기
                select_list_boxes = driver.find_elements(By.CLASS_NAME, "select-list-box")

                # 두 번째 div 태그의 HTML에서 a 태그들의 텍스트 값을 출력
                second_select_list_box = select_list_boxes[1]
                html_content = second_select_list_box.get_attribute("outerHTML")

                # BeautifulSoup을 사용하여 두 번째 div 태그의 HTML에서 a 태그들의 텍스트 값을 추출
                soup = BeautifulSoup(html_content, "html.parser")
                a_tags = soup.find_all("a")

                # 두 번째 div 태그의 a 태그들의 텍스트 값을 리스트에 저장
                second_a_texts = [tag.get_text(strip=True) for tag in a_tags]

                # 리스트에서 품절 여부를 판별
                for size_info in second_a_texts:
                    if second_word in size_info:
                        if '품절' in size_info:
                            print(f"{second_word}은(는) 품절입니다.")
                            soldout = "품절"
                        else:
                            print(f"{second_word}은(는) 품절이 아닙니다.")
                            soldout = ""
                        break
            else:
                print("본 제품은 1옵션 상품입니다.")
        else:
            soldout = "품절"
    else:
        soldout = ""

    print(f'현재 가격: {current_price} {soldout}')
    return current_price, soldout


def wizwid_logic(url, size_cell):
    # ,을 기준으로 문자열 분리
    split_values = str(size_cell).split(',')

    # 앞 뒤에 있는 단어를 각각 다른 변수에 저장
    first_word = split_values[0].strip() if len(split_values) > 0 else ""
    second_word = split_values[1].strip() if len(split_values) > 1 else ""

    # HTTP GET 요청
    response = requests.get(url, headers={"User-Agent": user_agent})

    # HTTP 요청이 성공했는지 확인
    if response.status_code == 200:
        # BeautifulSoup을 사용하여 HTML 파싱
        soup = BeautifulSoup(response.text, "html.parser")

        # price03 클래스를 가진 span 태그의 텍스트 출력
        price03_element = soup.find("span", class_="price03")
        if price03_element:
            current_price = price03_element.get_text(strip=True)
        else:
            print("price03 클래스를 가진 span 태그를 찾을 수 없습니다.")
            current_price = ""

        # optSelect0 name을 가진 select 태그의 하위 option 태그들의 텍스트 출력
        opt_select0_element = soup.find("select", {"name": "optSelect0"})
        if opt_select0_element:
            option_texts_opt0 = [option.get_text(strip=True) for option in opt_select0_element.find_all("option") if
                                 "품절" not in option.get_text(strip=True)]

            if first_word in option_texts_opt0:
                # optSelect1 name을 가진 select 태그의 하위 option 태그들의 텍스트 출력
                opt_select1_element = soup.find("select", {"name": "optSelect1"})
                if opt_select1_element:
                    driver.get(url)
                    for _ in range(3):
                        try:
                            optvalnm_option = wait.until(
                                EC.element_to_be_clickable((By.XPATH, f"//select[@name='optSelect0']")))
                            optvalnm_option.click()

                            optselect0_select = wait.until(
                                EC.element_to_be_clickable((By.XPATH, f"//option[@optvalnm='{first_word}']")))
                            optselect0_select.click()

                            # optvalnm이 "L"인 option 태그 클릭
                            optvalnm_option_2 = wait.until(
                                EC.element_to_be_clickable((By.XPATH, f"//select[@name='optSelect1']")))
                            optvalnm_option_2.click()

                            # 선택된 태그의 모든 하위 옵션 태그들의 텍스트 추출
                            options_text = [option.text for option in
                                            optvalnm_option_2.find_elements(By.TAG_NAME, 'option')
                                            if
                                            '품절' not in option.text]

                            # options_text 출력
                            print(options_text)
                            break
                        except StaleElementReferenceException:
                            continue
                    if second_word in options_text:
                        soldout = ""
                    else:
                        soldout = "품절"
                else:
                    print("본 제품은 단품 상품 또는 1개 옵션의 상품입니다.")
                    soldout = ""
            else:
                soldout = "품절"
                print("해당 제품의 옵션은 모두 품절되었습니다.")
        else:
            print("optSelect0 name을 가진 select 태그를 찾을 수 없습니다.")
            soldout = ""
    else:
        print(f"HTTP 요청 오류: {response.status_code}")
        current_price = ""
        soldout = ""

    print(f'현재 가격: {current_price} {soldout}')
    return current_price, soldout


# 전체 셀레니움
def galleria_logic(url, size_cell):
    # ,을 기준으로 문자열 분리
    split_values = str(size_cell).split(',')

    # 앞 뒤에 있는 단어를 각각 다른 변수에 저장
    first_word = split_values[0].strip() if len(split_values) > 0 else ""
    second_word = split_values[1].strip() if len(split_values) > 1 else ""

    # 웹 페이지 열기
    driver.get(url)

    # gds_amt 클래스를 가진 div 태그의 하위에서 dd 태그의 텍스트값 가져오기
    gds_amt_element = wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'gds_amt')))
    current_price = gds_amt_element.find_element(By.TAG_NAME, 'dd').text

    # gds_btns 클래스를 가진 div 태그 선택
    gds_btns_div = driver.find_element(By.CLASS_NAME, 'gds_btns')

    try:
        # disabled 클래스를 가진 button 태그 찾기
        disabled_button = gds_btns_div.find_element(By.CSS_SELECTOR, 'button[disabled="disabled"]')
        print("disabled 클래스를 가진 button이 있습니다.")
        soldout = "품절"
    except NoSuchElementException:
        print("disabled 클래스를 가진 button이 없습니다.")
        if str(first_word) != "None":
            # btn_opt_slt pg 클래스를 가진 첫 번째 div 태그 클릭
            first_btn_opt_slt = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, 'btn_opt_slt.pg')))
            first_btn_opt_slt.click()

            # 첫 번째 div 태그의 하위에서 option이라는 name을 가진 li 태그들 중에서 품절이 아닌 것들 추출
            option_list = first_btn_opt_slt.find_elements(By.CSS_SELECTOR, 'li[name="option"]')
            option_texts = [option.find_element(By.CSS_SELECTOR, 'span').text for option in option_list if
                            "품절" not in option.text]
            print(option_texts)

            # BKS가 리스트에 있다면 해당 값을 가진 li 태그 클릭
            if first_word in option_texts:
                bks_option = first_btn_opt_slt.find_element(By.CSS_SELECTOR, f'li[data-opt_val_nm="{first_word}"]')
                bks_option.click()
                soldout = ""
                if str(second_word) != "":
                    # 두 번째 div 태그가 존재한다면 클릭
                    second_btn_opt_slt = driver.find_elements(By.CLASS_NAME, 'btn_opt_slt.pg')[1]
                    if second_btn_opt_slt:
                        second_btn_opt_slt.click()

                        # 두 번째 div 태그의 하위에서 option이라는 name을 가진 li 태그들 중에서 품절이 아닌 것들 추출
                        second_option_list = second_btn_opt_slt.find_elements(By.CSS_SELECTOR, 'li[name="option"]')
                        second_option_texts = [option.find_element(By.CSS_SELECTOR, 'span').text for option in
                                               second_option_list if "품절" not in option.text]
                        print(second_option_texts)
                        if second_word in second_option_texts:
                            soldout = ""
                        else:
                            soldout = "품절"
                else:
                    pass
            else:
                soldout = "품절"
                print("본 제품의 해당 옵션은 품절입니다.")
        else:
            soldout = ""
            pass

    print(f'현재 가격: {current_price} {soldout}')
    return current_price, soldout


# 전체 셀레니움
def gsshop_logic(url, size_cell):
    # ,을 기준으로 문자열 분리
    split_values = str(size_cell).split(',')

    # 앞 뒤에 있는 단어를 각각 다른 변수에 저장
    first_word = split_values[0].strip() if len(split_values) > 0 else ""
    second_word = split_values[1].strip() if len(split_values) > 1 else ""

    # 웹 페이지 열기
    driver.get(url)

    # price-definition-ins 클래스를 가진 span 태그의 하위 태그들 중에서 strong 태그의 텍스트값 추출
    price_definition_ins = driver.find_element(By.CLASS_NAME, "price-definition-ins")
    strong_text = price_definition_ins.find_element(By.XPATH, ".//strong")
    current_price = strong_text.text

    try:
        prd_message_div = driver.find_element(By.CLASS_NAME, "prd_message.no-border")
        print("본 제품은 판매 종료된 상품입니다.")
        soldout = "품절"
    except NoSuchElementException:
        if first_word != "None":
            # sel 클래스를 가진 div 태그의 하위 태그들 중에서 첫 번째 태그 클릭
            sel = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "sel")))
            first_sel = sel.find_elements(By.XPATH, ".//*[contains(@id, 'Sel')]")[0]
            first_sel.click()

            # 첫 번째 태그의 하위 태그 중 ul 태그 찾기
            ul_element = first_sel.find_element(By.TAG_NAME, "ul")

            # ul_element의 텍스트값을 리스트에 저장
            ul_text = ul_element.text
            ul_text_list = ul_text.split('\n')
            print(ul_text_list)

            # 리스트에서 품절 여부를 판별
            for size_info in ul_text_list:
                if first_word in size_info:
                    if '품절' in size_info:
                        print(f"{first_word}은(는) 품절입니다.")
                        soldout = "품절"
                    else:
                        print(f"{first_word}은(는) 품절이 아닙니다.")
                        soldout = ""
                    break
            else:
                print(f"{first_word}은(는) 리스트에 없습니다.")
                soldout = "품절"

            if second_word != "":
                # DJ6914 블랙이 포함된 경우 해당 a 태그 클릭
                if first_word in ul_element.text:
                    a_tag_black = sel.find_element(By.XPATH, f".//a[contains(text(), '{first_word}')]")
                    a_tag_black.click()

                    # sel 클래스를 가진 div 태그의 두 번째 태그 클릭 (만약 두 번째 태그가 존재한다면)
                    second_sel = sel.find_elements(By.XPATH, ".//*[contains(@id, 'Sel')]")[2]
                    if second_sel:
                        second_sel.click()

                        # 두 번째 태그의 하위 태그 중 ul 태그 찾기
                        ul_element_second = second_sel.find_element(By.TAG_NAME, "ul")

                        # ul_element의 텍스트값을 리스트에 저장
                        ul_text = ul_element_second.text
                        ul_text_list_2 = ul_text.split('\n')
                        print(ul_text_list_2)
                    else:
                        print("본 제품은 단품 상품입니다.")
                        soldout = ""
                else:
                    print("해당 옵션은 모두 품절되었습니다.")
                    soldout = "품절"
            else:
                pass
        else:
            soldout = ""

    print(f'현재 가격: {current_price} {soldout}')
    return current_price, soldout


# 시트 이름과 사용자 정의 파싱 함수 매핑
sheet_custom_parsing_mapping = {
    'okmall': okmall_logic,
    'trendmecca': trendmecca_logic,
    'folder': folder_logic,
    'musinsa': musinsa_logic,
    'shoemarker': shoemarker_logic,
    'lotteon': lotteon_logic,
    'hmall': hmall_logic,
    'ssg': ssg_logic,
    'abcmart': abcmart_logic,
    'balaan': balaan_logic,
    'mustit': mustit_logic,
    'trenbe': trenbe_logic,
    'smarket': smarket_logic,
    'akmall': akmall_logic,
    # 'onthespot': onthespot_logic,
    'wconcept': wconcept_logic,
    'wizwid': wizwid_logic,
    'galleria': galleria_logic,
    'gsshop': gsshop_logic,
}

# 반복문 실행
for iteration in range(repeat):
    print(f"{iteration + 1}번째 반복 시작")

    # 엑셀 파일 열기
    workbook = openpyxl.load_workbook('discount_data.xlsx')

    # 모든 시트를 순회하면서 함수 호출
    for sheet_name, custom_parsing_func in sheet_custom_parsing_mapping.items():
        current_sheet = workbook[sheet_name]
        process_sheet(current_sheet, custom_parsing_func)

    # 엑셀 파일 저장
    workbook.save('discount_data.xlsx')

    print(f"{iteration + 1}번째 반복 완료 및 엑셀 파일 저장")

print("저장되었습니다.")

driver.quit()
